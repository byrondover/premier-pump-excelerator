#!/usr/bin/env python

from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")

from collections import OrderedDict

from openpyxl import load_workbook
from openpyxl.styles import Font

bold = Font(bold=True)

wb = load_workbook('test.xlsx')
print wb.get_sheet_names()

weld_picklist_sheet = wb.create_sheet()
weld_picklist_sheet.title = 'WELD SCF Picklist'
print wb.get_sheet_names()

parts_list_sheet = wb.active

def find_first_row(row = 0):
    cell_value = str()

    while str(cell_value) != 'QTY':
        cell_value = parts_list_sheet.rows[row][0].value
        row += 1

    return row

def find_last_row(row):
    cell_value = str()

    while cell_value != None:
        cell_value = parts_list_sheet.rows[row][0].value
        row += 1

    return row - 1

start_row_number = find_first_row()
start_row = 'A' + str(start_row_number)
end_row_number = find_last_row(start_row_number)
end_row = 'Z' + str(end_row_number)

#fabricated_parts = parts_list[start_row:end_row]
fabricated_parts = list(parts_list_sheet.iter_rows(start_row + ':' + end_row))

print '\n\n\n===FABRICATED PARTS===\n\n\n'
for row in fabricated_parts:
    for cell in row:
        print cell.value,
    print('\n')

#import pdb
#pdb.set_trace()

start_row_number = find_first_row(end_row_number + 1)
start_row = 'A' + str(start_row_number)
end_row_number = find_last_row(start_row_number + 1)
end_row = 'Z' + str(end_row_number)

#weldments = parts_list[start_row:end_row]
weldments = list(parts_list_sheet.iter_rows(start_row + ':' + end_row))

print '\n\n\n===WELDMENTS===\n\n\n'
for row in weldments:
    for cell in row:
        print cell.value,
    print('\n')

start_row_number = find_first_row(end_row_number + 1)
start_row = 'A' + str(start_row_number)
end_row_number = find_last_row(start_row_number + 1)
end_row = 'Z' + str(end_row_number)

#purchased_parts = parts_list[start_row:end_row]
purchased_parts = list(parts_list_sheet.iter_rows(start_row + ':' + end_row))

print '\n\n\n===PURCHASED===\n\n\n'
for row in purchased_parts:
    for cell in row:
        print cell.value,
    print('\n')

#cols = parts_list.get_highest_column()
#headers = dict( (i, parts_list.cell(row=56, column=i).value) for i in range(cols) )

def create_section_list(section):
    headers = [i.value for i in fabricated_parts[0] if i.value != None]
    print headers

    def item(i, j):
        return (fabricated_parts[0][j].value, section[i][j].value)

    return [OrderedDict(item(i, j) for j in range(len(headers))) for i in range(1,
        len(section))]
    #return [dict(item(i, j) for j in range(len(headers))) for i in range(1,
    #    len(section))]

fabricated_list = create_section_list(fabricated_parts)
weldments_list = create_section_list(weldments)
purchased_list = create_section_list(purchased_parts)

print purchased_list

full_list = fabricated_list + weldments_list + purchased_list

#import pdb
#pdb.set_trace()

for i, header in enumerate(fabricated_list):
    if str(header) == 'WELDED':
        welded_idx = i
    if str(header) == 'WELDMENT USED':
        weldment_used_idx = i

def append_data(sheet, data):
    sheet.append(data[0].keys())
    for row in data:
        sheet.append(row.values())

weld_picklist_data = sorted(fabricated_list, key=lambda k: k['PART NUMBER'])
append_data(weld_picklist_sheet, weld_picklist_data)

weld_bom_sheet = wb.create_sheet()
weld_bom_sheet.title = 'WELD BOM'

weld_bom = [x for x in full_list if str(x['WELDED']) == 'WELDED' and
    str(x['WELDMENT USED']) != 'SHIPPED LOOSE']
weld_bom_data = sorted(weld_bom, key=lambda k: k['PART NUMBER'])

append_data(weld_bom_sheet, weld_bom_data)

weld_loose_sheet = wb.create_sheet()
weld_loose_sheet.title = 'WELD LOOSE'

weld_loose = [x for x in full_list if str(x['WELDED']) == 'WELDED' and
    str(x['WELDMENT USED']) == 'SHIPPED LOOSE']
weld_loose_data = sorted(weld_loose, key=lambda k: k['PART NUMBER'])

append_data(weld_loose_sheet, weld_loose_data)

weld_packing_sheet = wb.create_sheet()
weld_packing_sheet.title = 'WELD Packing Slip'

weld_packing = [x for x in full_list if str(x['WELDMENT USED']) ==
        'SHIPPED LOOSE']
weld_packing_data = sorted(weld_packing, key=lambda k: k['PART NUMBER'])

append_data(weld_packing_sheet, weld_packing_data)

finish_picklist_sheet = wb.create_sheet()
finish_picklist_sheet.title = 'FINISH Pick List'

finish_picklist = [x for x in full_list if str(x['WELDMENT USED']) ==
        'SHIPPED LOOSE']
finish_picklist_data = sorted(finish_picklist, key=lambda k: k['PART NUMBER'])

append_data(finish_picklist_sheet, finish_picklist_data)

#import pdb
#pdb.set_trace()

# Apply styles.
for i in range(1, len(wb.worksheets)):
    dims = {}
    ws = wb.worksheets[i]

    for cell in ws.rows[0]:
        cell.font = bold

    for row in ws.rows:
        for cell in row:
            if cell.value:
                #foo = max((dims.get(cell.column, 0), len(str(cell.value))))
                #print "cell value: " + str(cell.value)
                #print "max: " + str(foo)
                #print "dims: " + str(dims.get(cell.column, 0))
                #print "len: " + str(len(str(cell.value)))
                dims[cell.column] = max((dims.get(cell.column, 0),
                    len(str(cell.value)), 4))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

#for col in headers:
#    for row in

#while found != True:
#    i += 1
#    for j in range(0, 10):
#        cell_value = parts_list.rows[i][j].value
#        print cell_value
#        if str(cell_value) == 'FABRICATED PARTS':
#            found = True

#print cell_value

#import pdb
#pdb.set_trace()

wb.save("test_complete.xlsx")

#import xlrd
#book = xlrd.open_workbook("test.xls")
#print("The number of worksheets is {0}".format(book.nsheets))
#print("Worksheet name(s): {0}".format(book.sheet_names()))
#sh = book.sheet_by_index(0)
#print("{0} {1} {2}".format(sh.name, sh.nrows, sh.ncols))
#print("Cell D30 is {0}".format(sh.cell_value(rowx=29, colx=3)))
#for rx in range(sh.nrows):
#  print(sh.row(rx))

#from xlwt import Workbook

#book = Workbook()

#sheet1 = book.add_sheet('PARTS LIST')
