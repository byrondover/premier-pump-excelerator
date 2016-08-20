#!/usr/bin/env python

from collections import OrderedDict

from openpyxl import load_workbook
from openpyxl.styles import Font

def append_data(sheet, data):
    sorted_data = sorted(data, key=lambda k: k['PART NUMBER'])

    sheet.append(sorted_data[0].keys())

    for row in data:
        sheet.append(row.values())

def create_parts_list(sheet, initial_row = 0):
    first_row_number = find_first_row(sheet, initial_row)
    first_row = 'A' + str(first_row_number)

    last_row_number = find_last_row(sheet, first_row_number)
    last_row = 'Z' + str(last_row_number)

    parts_list = list(sheet.iter_rows(first_row + ':' + last_row))

    return parts_list, last_row_number

def create_section_list(source_cells, section):
    headers = [i.value for i in source_cells[0] if i.value != None]

    def item(i, j):
        return (source_cells[0][j].value, section[i][j].value)

    return [OrderedDict(item(i, j) for j in range(len(headers)))
        for i in range(1, len(section))]

def create_sheet(workbook, name):
    sheet = workbook.create_sheet()
    sheet.title = name

    return sheet

def find_first_row(source_sheet, row = 0):
    cell_value = str()

    while str(cell_value) != 'QTY':
        cell_value = source_sheet.rows[row][0].value
        row += 1

    return row

def find_last_row(source_sheet, row):
    cell_value = str()
    print row

    while cell_value != None:
        cell_value = source_sheet.rows[row][0].value
        row += 1

    return row - 1

def excelerate(file):
    # Load spreadsheet into Workbook object.
    wb = load_workbook(file)

    # Set styles variables for later use.
    bold = Font(bold=True)

    weld_picklist_sheet = wb.create_sheet()
    weld_picklist_sheet.title = 'WELD SCF Picklist'

    parts_list_sheet = wb.active

    start_row_number = find_first_row(parts_list_sheet)
    start_row = 'A' + str(start_row_number)
    end_row_number = find_last_row(parts_list_sheet, start_row_number)
    end_row = 'Z' + str(end_row_number)

    #fabricated_parts = parts_list[start_row:end_row]
    fabricated_parts = list(parts_list_sheet.iter_rows(start_row + ':' + end_row))

    start_row_number = find_first_row(parts_list_sheet, end_row_number + 1)
    start_row = 'A' + str(start_row_number)
    end_row_number = find_last_row(parts_list_sheet, start_row_number + 1)
    end_row = 'Z' + str(end_row_number)

    #weldments = parts_list[start_row:end_row]
    weldments = list(parts_list_sheet.iter_rows(start_row + ':' + end_row))

    start_row_number = find_first_row(parts_list_sheet, end_row_number + 1)
    start_row = 'A' + str(start_row_number)
    end_row_number = find_last_row(parts_list_sheet, start_row_number + 1)
    end_row = 'Z' + str(end_row_number)

    #purchased_parts = parts_list[start_row:end_row]
    #purchased_parts = list(parts_list_sheet.iter_rows(start_row + ':' + end_row))
    fabricated_parts, last_row_number = create_parts_list(parts_list_sheet)

    weldments, last_row_number = create_parts_list(
        parts_list_sheet, last_row_number + 1)

    purchased_parts, last_row_number = create_parts_list(
        parts_list_sheet, last_row_number + 1)

    fabricated_list = create_section_list(fabricated_parts, fabricated_parts)
    weldments_list = create_section_list(fabricated_parts, weldments)
    purchased_list = create_section_list(fabricated_parts, purchased_parts)

    full_list = fabricated_list + weldments_list + purchased_list

    """
    for i, header in enumerate(fabricated_list):
        if str(header) == 'WELDED':
            welded_idx = i
        if str(header) == 'WELDMENT USED':
            weldment_used_idx = i
    """

    append_data(weld_picklist_sheet, fabricated_list)

    weld_bom_sheet = create_sheet(wb, 'WELD BOM')

    weld_bom_data = [x for x in full_list if str(x['WELDED']) == 'WELDED' and
        str(x['WELDMENT USED']) != 'SHIPPED LOOSE']

    append_data(weld_bom_sheet, weld_bom_data)

    weld_loose_sheet = create_sheet(wb, 'WELD LOOSE')

    weld_loose_data = [x for x in full_list if str(x['WELDED']) == 'WELDED' and
        str(x['WELDMENT USED']) == 'SHIPPED LOOSE']

    append_data(weld_loose_sheet, weld_loose_data)

    weld_packing_sheet = create_sheet(wb, 'WELD Packing Slip')

    weld_packing_data = [x for x in full_list if str(x['WELDMENT USED']) ==
            'SHIPPED LOOSE']

    append_data(weld_packing_sheet, weld_packing_data)

    finish_picklist_sheet = create_sheet(wb, 'FINISH Pick List')

    finish_picklist_data = [x for x in full_list if str(x['WELDMENT USED']) ==
            'SHIPPED LOOSE']

    append_data(finish_picklist_sheet, finish_picklist_data)

    # Apply styles.
    for i in range(1, len(wb.worksheets)):
        dims = {}
        ws = wb.worksheets[i]

        for cell in ws.rows[0]:
            cell.font = bold

        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0),
                        len(str(cell.value)), 4))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value

    wb.save("test_complete.xlsx")

excelerate('test.xlsx')
