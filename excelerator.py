import copy
import os
import re
from collections import OrderedDict
from datetime import datetime
from io import BytesIO

import xlrd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.writer.excel import save_virtual_workbook


class Excelerator:

    def __init__(self, workbook=str(), multiplier=5, order_number=str(),
            primary_color=str(), secondary_color=str()):
        self.filename = workbook
        self.multiplier = multiplier
        self.primary_color = primary_color.strip()
        self.secondary_color = secondary_color.strip()

        if not isinstance(self.filename, str):
            self.filename = workbook.filename

        self.filename_stripped, self.extension = os.path.splitext(self.filename)

        if order_number:
            self.order_number = order_number
        else:
            self.order_number = self.sanitize_name(self.filename_stripped,
                                                   max_length=64)

        # Set style variables for later use
        self.side = Side(border_style='thin')

        self.bold = Font(bold=True)
        self.border = Border(left=self.side, right=self.side,
                             top=self.side, bottom=self.side)
        self.date_format = 'mm/dd/yy'
        self.fill = PatternFill(patternType='solid', fgColor='D9D9D9')
        self.title_font = Font(size=24)

        # Excelerate immediately when parsable file provided
        if self.filename:
            self.excelerate(workbook)

    def add_column(self, name, parts_list, last=False):
        for part in parts_list:
            part.update({name: None})
            part.move_to_end(name, last=last)

    def append_colors_legend(self, sheet, columns=4, pad=True):
        rows_added = 0

        # Don't bother if not colors are provided
        if not self.primary_color and not self.secondary_color:
            return rows_added

        if self.primary_color:
            p_value = 'PRIMARY COLOR :        {}'.format(self.primary_color)
            sheet.append([None, p_value])
            rows_added += 1

            p_row = sheet.max_row
            p_cell = sheet.cell(column=2, row=p_row)
            p_cell.font = self.bold
            sheet.row_dimensions[p_row].height = 30
            sheet.merge_cells(start_row=p_row, start_column=2,
                              end_row=p_row, end_column=columns)

        if self.secondary_color:
            s_value = 'SECONDARY COLOR :   {}'.format(self.secondary_color)
            sheet.append([None, s_value])
            rows_added += 1

            s_row = sheet.max_row
            s_cell = sheet.cell(column=2, row=s_row)
            s_cell.font = self.bold
            sheet.row_dimensions[s_row].height = 30
            sheet.merge_cells(start_row=s_row, start_column=2,
                              end_row=s_row, end_column=columns)

        if pad:
            # Pad colors legend with blank rows
            self.append_empty_row(sheet, 2)
            rows_added += 2

        return rows_added

    def append_data(self, data, sheet):
        sorted_data = sorted(data, key=lambda k: str(k.get('PART NUMBER')))

        # Append dictionary keys as spreadsheet headers.
        sheet.append(list(sorted_data[0]))

        for row in sorted_data:
            sheet.append(list(row.values()))

    def append_empty_row(self, sheet, number=1):
        for i in range(number):
            sheet.append([str()])

    def append_signature(self, prompt, sheet, columns=5, date=True):
        signature_line = ': ___________________________'
        signature = str(prompt) + signature_line

        if date:
            signature += '                    ' + 'Date' + signature_line

        self.append_empty_row(sheet, 2)
        sheet.append([None, signature])

        # Style signature cells.
        signature_row = sheet.max_row
        sheet.merge_cells(start_row=signature_row, start_column=2,
                          end_row=signature_row, end_column=columns)

    def append_title(self, sheet, title=None, columns=4, pad=True):
        if not title:
            title_components = [
                'SO {}'.format(self.order_number.upper()),
                sheet.title,
                'QTY {n}'.format(n=self.multiplier)
            ]
            title = ' – '.join(title_components)

        sheet.append([title])

        # Style sheet title.
        title_row = sheet.max_row
        sheet.merge_cells(start_row=title_row, start_column=1,
                          end_row=title_row, end_column=columns)
        sheet.row_dimensions[title_row].height = 30
        title_cell = sheet.cell(column=1, row=title_row)
        title_cell.alignment = Alignment(horizontal='center')
        title_cell.font = self.title_font

        # Pad title with blank rows.
        if pad:
            self.append_empty_row(sheet, 2)

    def apply_styles(self, sheet, start_row=3):
        dims = dict()
        rows = list(sheet.rows)

        for cell in rows[start_row]:
            cell.border = self.border
            cell.font = self.bold

        for i in range(start_row, len(rows)):
            padding = 2 if i == 0 else 1
            sheet.row_dimensions[i + 1].height = 30

            for cell in rows[i]:
                cell.border = self.border

                # Shade every other data row
                if (i + start_row) % 2 == 1:
                    cell.fill = self.fill

                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0),
                        len(str(cell.value)) + padding, 4))

                    if isinstance(cell.value, datetime):
                        dims[cell.column] = 10
                        cell.number_format = self.date_format

        for col, value in dims.items():
            sheet.column_dimensions[col].width = value

    def convert_to_xlsx(self, file_):
        # Read xls using xlrd
        if isinstance(file_, str):
            book = xlrd.open_workbook(file_)
        else:
            book = xlrd.open_workbook(file_contents=file_.stream.read())

        index = 0
        nrows, ncols = 0, 0

        while nrows * ncols == 0:
            sheet = book.sheet_by_index(index)
            nrows = sheet.nrows
            ncols = sheet.ncols
            index += 1

        # Prepare an xlsx workbook
        workbook = Workbook()
        worksheet = workbook.active

        for row in range(0, nrows):
            for col in range(0, ncols):
                worksheet.cell(
                    row=row + 1, column=col + 1).value = sheet.cell_value(
                        row, col)

        return workbook

    def create_headers_list(self, source_cells):
        row_number = str(self.find_first_row() + 1)
        headers_generator = self.master_parts_sheet[
            'A' + row_number + ':Z' + row_number]
        headers_list = headers_generator[0]

        return headers_list

    def create_parts_list(self, initial_row=0):
        first_row_number = self.find_first_row(initial_row)
        last_row_number = self.find_last_row(first_row_number)

        first_row = 'A' + str(first_row_number + 1)
        last_row = 'Z' + str(last_row_number + 1)

        parts_list = list(self.master_parts_sheet[first_row + ':' + last_row])

        return parts_list, last_row_number

    def create_section_list(self, section, part_group):
        headers = [i.value for i in self.headers if i.value != None]

        def item(i, j):
            value = section[i][j].value

            if self.headers[j].value == 'QTY':
                value = int(value) * self.multiplier

            if self.headers[j].value == 'PART NUMBER':
                if isinstance(value, float):
                    value = '{:.0f}'.format(value)
                else:
                    value = str(value)

            if j == len(headers):
                attr = ('PART GROUP', part_group)
            else:
                attr = (self.headers[j].value, value)

            return attr

        return [OrderedDict(item(i, j) for j in range(len(headers) + 1))
            for i in range(1, len(section))]

    def create_sheet(self, name):
        sheet = self.workbook.create_sheet()
        sanitized_name = self.sanitize_name(name)

        try:
            # Try to name the sheet...
            sheet.title = sanitized_name
        except ValueError:
            # ...but if we can't, no big deal, just stick with the default
            pass

        return sheet

    def create_sheet__generic(self, name, section, columns, sort,
                              filter_='True', secondary_sort=None,
                              colors_legend=False):
        column_map = dict()
        headers = [i.value for i in self.headers if i.value != None]
        parts = list()
        start_row = 3

        # Generate column header map
        for value in columns.values():
            if value:
                column_map[value] = headers.index(value)

        for i in range(1, len(section)):
            cell_map = dict()

            for j in range(len(headers)):
                header = headers[j]
                value = section[i][j].value

                if header == 'QTY':
                    value = int(value) * self.multiplier

                if header == 'PART NUMBER':
                    if isinstance(value, float):
                        value = '{:.0f}'.format(value)
                    else:
                        value = str(value)

                cell_map[header] = value

            if not (eval(filter_)):
                continue

            row_data = OrderedDict()

            for custom_header, key in columns.items():
                index = column_map.get(key)
                value = cell_map[key] if key != None else None

                row_data[custom_header] = value

            parts.append(row_data)

        # Skip sheet creation if no parts match
        if not len(parts):
            return

        sheet = self.create_sheet(name)

        self.append_title(sheet, columns=len(columns))

        if colors_legend:
            modifier = self.append_colors_legend(sheet,
                                                 columns=len(columns) - 1)
            start_row += modifier

        sorted_parts = copy.deepcopy(parts)
        if secondary_sort:
            sorted_parts = sorted(
                sorted_parts, key=lambda k: str(k.get(secondary_sort)))
        sorted_parts = sorted(sorted_parts, key=lambda k: str(k.get(sort)))

        # Append dictionary keys as spreadsheet headers
        sheet.append(list(sorted_parts[0]))

        for row in sorted_parts:
            sheet.append(list(row.values()))

        self.apply_styles(sheet, start_row=start_row)
        self.append_signature('Received by', sheet, columns=len(columns) - 1)

        # Lastly, configure page setup and printable area
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToPage = True
        sheet.page_setup.fitToHeight = False

    def create_sheet__generic_weldments(self, section, columns, sort):
        column_map = dict()
        headers = [i.value for i in self.headers if i.value != None]
        parts_by_weldment = dict()

        # Generate column header map
        for value in columns.values():
            if value:
                column_map[value] = headers.index(value)

        for i in range(1, len(section)):
            cell_map = dict()
            weldment = None

            for j in range(len(headers)):
                header = headers[j]
                value = section[i][j].value

                if header == 'QTY':
                    value = int(value) * self.multiplier

                if header == 'PART NUMBER':
                    if isinstance(value, float):
                        value = '{:.0f}'.format(value)
                    else:
                        value = str(value)

                if header == 'WELDMENT USED':
                    if value and str(value).strip() != 'SHIPPED LOOSE':
                        weldment = str(value).strip()

                cell_map[header] = value

            row_data = OrderedDict()

            for custom_header, key in columns.items():
                index = column_map.get(key)
                value = cell_map[key] if key != None else None

                row_data[custom_header] = value

            if weldment:
                parts_by_weldment.setdefault(weldment, []).append(row_data)

        for weldment in sorted(parts_by_weldment):
            sheet = self.create_sheet(weldment)
            title = 'SO {} – {} – {}'.format(self.order_number.upper(),
                                             sheet.title,
                                             'QTY {n}'.format(n=self.multiplier))

            self.append_title(sheet, 'Weldment', columns=len(columns),
                              pad=False)
            self.append_title(sheet, title, columns=len(columns))

            parts = copy.deepcopy(parts_by_weldment[weldment])
            sorted_parts = sorted(parts,
                                  key=lambda k: str(k.get('PART_NUMBER')))

            # Append dictionary keys as spreadsheet headers
            sheet.append(list(sorted_parts[0]))

            for row in sorted_parts:
                sheet.append(list(row.values()))

            self.apply_styles(sheet, start_row=4)
            self.append_signature('Received by', sheet,
                                  columns=len(columns) - 1)

            # Lastly, configure page setup and printable area
            sheet.page_setup.orientation = "landscape"
            sheet.page_setup.fitToPage = True
            sheet.page_setup.fitToHeight = False

    def create_sheet_bend(self, section):
        columns = OrderedDict([
            ('QTY NEEDED', 'QTY'),
            ('QTY RCD', None),
            ('PART NUMBER', 'PART NUMBER'),
            ('DESCRIPTION', 'DESCRIPTION'),
            ('MATERIAL', 'MATERIAL'),
            ('REV', 'REV'),
            ('LAST REV', 'LAST REV'),
            ('WELDED', 'WELDED'),
            ('WELDMENT USED', 'WELDMENT USED'),
            ('PROGRAM TIME IN', None),
            ('TIME IN', None),
            ('TIME OUT', None),
            ('TOTAL', None)
        ])
        filter_ = "'FORMED' in str(cell_map['PROCESS'])"
        sort = 'MATERIAL'
        secondary_sort = 'PART NUMBER'

        self.create_sheet__generic('Bend', section, columns, sort, filter_,
                                  secondary_sort)

    def create_sheet_finish_slip(self, section):
        columns = OrderedDict([
            ('DLVD', None),
            ('RCVD', None),
            ('QTY', 'QTY'),
            ('PART NUMBER', 'PART NUMBER'),
            ('DESCRIPTION', 'DESCRIPTION'),
            ('MATERIAL', 'MATERIAL'),
            ('WELDED', 'WELDED'),
            ('COLOR', 'COLOR')
        ])
        filter_ = "str(cell_map['WELDMENT USED'].strip()) == 'SHIPPED LOOSE'"
        sort = 'COLOR'
        secondary_sort = 'PART NUMBER'

        self.create_sheet__generic('Finish Pack Slip', section, columns, sort,
                                   filter_, secondary_sort, colors_legend=True)

    def create_sheet_job_inventory(self, section):
        columns = OrderedDict([
            ('QTY NEEDED', 'QTY'),
            ('QTY IN STOCK', None),
            ('QTY RCD', None),
            ('PART NUMBER', 'PART NUMBER'),
            ('DESCRIPTION', 'DESCRIPTION'),
            ('MATERIAL', 'MATERIAL'),
            ('PROCESS', 'PROCESS')
        ])
        sort = 'MATERIAL'
        secondary_sort = 'PART NUMBER'

        self.create_sheet__generic('Job Inventory', section, columns, sort,
                                   secondary_sort=secondary_sort)

    def create_sheet_weld_pack_slip(self, section):
        columns = OrderedDict([
            ('DLVD', None),
            ('RCVD', None),
            ('QTY', 'QTY'),
            ('PART NUMBER', 'PART NUMBER'),
            ('DESCRIPTION', 'DESCRIPTION'),
            ('MATERIAL', 'MATERIAL'),
            ('WELDED', 'WELDED'),
            ('COLOR', 'COLOR')
        ])
        filter_ = "str(cell_map['WELDMENT USED'].strip()) == 'SHIPPED LOOSE'"
        sort = 'COLOR'
        secondary_sort = 'PART NUMBER'

        self.create_sheet__generic('Weld Pack Slip', section, columns, sort,
                                   filter_, secondary_sort, colors_legend=True)

    def create_sheet_weld_pick_list(self, section):
        columns = OrderedDict([
            ('QTY NEEDED', 'QTY'),
            ('DLVD', None),
            ('PALLET', None),
            ('RCVD', None),
            ('PART NUMBER', 'PART NUMBER'),
            ('DESCRIPTION', 'DESCRIPTION'),
            ('MATERIAL', 'MATERIAL'),
            ('REV', 'REV'),
            ('LAST REV', 'LAST REV'),
            ('WELDED', 'WELDED'),
            ('WELDMENT USED', 'WELDMENT USED')
        ])
        filter_ = "str(cell_map['WELDED'].strip()) == 'WELDED'"
        sort = 'WELDMENT USED'
        secondary_sort = 'PART NUMBER'

        self.create_sheet__generic('Weld Pick List', section, columns, sort,
                                   filter_, secondary_sort)

    def create_sheets_weldments(self, section):
        columns = OrderedDict([
            ('QTY NEEDED', 'QTY'),
            ('DLVD', None),
            ('RCVD', None),
            ('PART NUMBER', 'PART NUMBER'),
            ('DESCRIPTION', 'DESCRIPTION'),
            ('MATERIAL', 'MATERIAL'),
            ('REV', 'REV'),
            ('LAST REV', 'LAST REV'),
            ('WELDED', 'WELDED'),
            ('WELDMENT USED', 'WELDMENT USED')
        ])
        sort = 'PART NUMBER'

        self.create_sheet__generic_weldments(section, columns, sort)

    def find_first_row(self, row=0):
        cell_value = str()

        while str(cell_value) != 'QTY':
            row += 1

            if row == self.master_parts_sheet.max_row:
                break

            cell_value = self.master_parts_sheet.cell(row=row, column=1).value

        return row - 1

    def find_last_row(self, row):
        cell_value = True

        while cell_value:
            row += 1

            if row == self.master_parts_sheet.max_row + 1:
                break

            cell_value = self.master_parts_sheet.cell(row=row, column=1).value

        return row - 2

    def get_workbook(self):
        workbook = None

        if hasattr(self, 'workbook'):
            workbook = self.workbook

        return workbook

    def get_workbook_stream(self):
        workbook = None

        if hasattr(self, 'workbook'):
            wb_bytes = save_virtual_workbook(self.workbook)
            workbook = BytesIO(wb_bytes)

        return workbook

    def sanitize_name(self, name=str(), max_length=30):
        sanitized_name = re.sub(r'[\\*?:/\[\]]', str(), name)
        valid_name = re.sub(r' +', ' ', sanitized_name)[:max_length].strip()
        return valid_name

    def excelerate(self, file_):
        # Limit base filename to 64 characters
        file_.filename = self.sanitize_name(file_.filename, max_length=64)

        # Load spreadsheet into Workbook object
        if self.extension == '.xlsx':
            self.workbook = load_workbook(file_)
        else:
            self.workbook = self.convert_to_xlsx(file_)

        # First spreadsheet should contain master parts list
        self.master_parts_sheet = self.workbook.worksheets[0]

        # Iterate through master parts list and identify each section
        fabricated_parts, last_row_number = self.create_parts_list()
        weldments, last_row_number = self.create_parts_list(
            last_row_number + 1)
        purchased_parts, last_row_number = self.create_parts_list(
            last_row_number + 1)

        master_parts = fabricated_parts + weldments[1:] + purchased_parts[1:]

        # Create master list of headers
        self.headers = self.create_headers_list(master_parts)

        # Create lists of dictionarys for each section
        fabricated_list = self.create_section_list(
            fabricated_parts, 'FabricatedParts')
        weldments_list = self.create_section_list(
            weldments, 'WeldmentParts')
        purchased_list = self.create_section_list(
            purchased_parts, 'PurchasedFabParts')

        master_parts_list = fabricated_list + weldments_list + purchased_list

        self.create_sheet_job_inventory(fabricated_parts)
        self.create_sheet_bend(fabricated_parts)
        self.create_sheet_weld_pick_list(fabricated_parts)
        self.create_sheet_weld_pack_slip(fabricated_parts + weldments[1:])
        self.create_sheet_finish_slip(fabricated_parts + weldments[1:])
        self.create_sheets_weldments(fabricated_parts + purchased_parts[1:])

        return self.workbook
